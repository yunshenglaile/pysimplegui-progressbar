import openpyxl
import PySimpleGUI as sg
import time

wb=openpyxl.load_workbook('12-3.xlsx')
ws=wb['优秀员工表']
num=0
if not '优秀员工表3' in wb.sheetnames:
    nws = wb.create_sheet('优秀员工表3')
nws.append(['序号','部门','姓名'])
progressbar = [ [sg.ProgressBar(len(list(ws.iter_rows(min_row=2))), orientation= 'h' , size=(51, 10), key= '-progressbar-' )]]
outputwin = [ [sg.Output(size=(78,20))]]
layout = [ [sg.Frame('Progress',layout= progressbar,key='-Progress-')], [sg.Frame( 'Output' , layout = outputwin,key='-Output-')], [sg.Submit( '执行'),sg.Cancel('取消')]]
window = sg.Window( '自定义进度条', layout)
progress_bar = window[ '-progressbar-' ]

while True:
  event, values = window.read(timeout=10)
  # print(event)
  if event ==  '取消'  or event is None:
    break
  elif event ==  '执行' :
    for i,item in enumerate(list(ws.iter_rows(min_row=2))):
        for val in item[1].value.split('、'):
            num +=1
            nws.append((num,item[0].value,val))
            print("第" + str(i+ 1) + "行处理完成！")
            time.sleep(0.05)
            progress_bar.UpdateBar(i + 1)
    wb.save('12-3test.xlsx')
